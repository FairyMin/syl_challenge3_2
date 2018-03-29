# -*- coding:utf-8 -*-

from openpyxl import load_workbook #可以用来载入已有数据表格

from openpyxl import Workbook #可以用来处理新的数据表格

import datetime #可以用来处理时间相关的数据

def combine():
    '''
    该函数可以用来处理原数据文件：
    1. 合并表格写入的combine表中
    2.保存原数据文件
    '''
    #载入excel文件
    wb1 = load_workbook('courses.xlsx')
    #获取excel文件中的所有sheet名的列表
    sheet_list = wb1.get_sheet_names()
    #读取名为students和time的 sheet 页
    sheet1 = wb1['students']
    sheet2 = wb1['time']
    #如果文当中没有combine这个sheet，则创建
    if 'combine' not in sheet_list:
        wb1.create_sheet('combine',index=2)
    #读取sheet   combine
    sheet3 = wb1['combine']
    #读取最大行数，最大列数
    max_row_s1 = sheet1.max_row
    max_column_s1 = sheet1.max_column

    max_row_s2 = sheet2.max_row
    max_column_s2 = sheet2.max_column
    
    #创建一个字典存储sheet1中的数据,
    #格式为： ｛'课程名称':{'创建时间':'time','学习人数':'num'}｝
    sheet1_dic = {}
    for i in range(2,max_row_s1+1):
        sheet1_dic.setdefault(sheet1.cell(row=i,column=2).value,{'cre_time':'','num':''})
        sheet1_dic[sheet1.cell(row=i,column=2).value]['cre_time'] = sheet1.cell(row=i,column=1).value
        sheet1_dic[sheet1.cell(row=i,column=2).value]['num'] = sheet1.cell(row=i,column=3).value
        
    #创建一个字典存储sheet2中的数据
    #格式为：｛'课程名称':{"学习时间":time}｝
    sheet2_dic = {}
    for j in range(2,max_row_s2+1):
        sheet2_dic.setdefault(sheet2.cell(row=j,column=2).value,{'learn_time':0})
        sheet2_dic[sheet2.cell(row=j,column=2).value]['learn_time'] = sheet2.cell(row=j,column=3).value
    

    row_title = ['创建时间','课程名称','学习人数','学习时间']
    sheet3.append(row_title)
    
    line_num=1
    for key in sheet1_dic:
        line_num+=1
        sheet3.cell(row=line_num,column=2).value = key
        sheet3.cell(row=line_num,column=1).value = sheet1_dic[key]['cre_time']
        sheet3.cell(row=line_num,column=3).value = sheet1_dic[key]['num']
        if key in sheet2_dic:
            sheet3.cell(row=line_num,column=4).value = sheet2_dic[key]['learn_time']

    wb1.save('/home/shiyanlou/Code/courses.xlsx')
    
#    for i in range(1,max_row_s1+1):
#        for j in range(1,max_column_s1+1):  #chr(97)='a'
#            sheet3.cell(row=i,column=j).value = sheet1.cell(row=i,column=j).value
    


    #test：获取1，10行中的第2列数据
#    li=[]
#    for row_num in range(1,10):
#        li.append(sheet1.cell(row=row_num,column=2).value)
#    print(li)

    


def split():
    '''
    该函数可以用来分割文件:
    1.读取combine表中的数据
    2.将数据按时间分割
    3.写入不同的数据表中
    '''
    #读取sheet   combine
    wb = load_workbook('/home/shiyanlou/Code/courses.xlsx')
    sheet3 = wb['combine']

    #获取sheet combine的最大行数于列数
    max_row_s3 = sheet3.max_row
    max_column_s3 = sheet3.max_column
    
    #将读取的数据存储在字典里
    #格式为：{'年份'：[{'课程名称':"name",'学习人数':num,'学习时间':learn_time},{},..]}
    info_dic = {}
    year_set = set()
    for i in range(2,max_row_s3+1):
        x_a = sheet3.cell(row=i,column=1).value
        year_set.add(x_a.year)
        if x_a.year not in info_dic.keys():
            info_dic[x_a.year] = []
        for j in year_set:
            if j == x_a.year:
                info_dic[j].append({"c_name":sheet3.cell(row=i,column=2).value,
                        "num":sheet3.cell(row=i,column=3).value,
                        "learn_time":sheet3.cell(row=i,column=4).value,
                        "cre_time":sheet3.cell(row=i,column=1).value
                        })
    wb.close()
    #print(info_dic[2013])
    #将数据存入excel文件
    for year in info_dic.keys():
        wb_new = Workbook()
        #wb_new.save('/home/shiyanlou/Code/%s.xlsx'%year
        wb_new.create_sheet('%s'%year,index=0)
        sname = wb_new.get_sheet_names()
        print(sname)
        sheet_new = wb_new['%s'%str(year)]
        row_new=1
        row_title = ['创建时间','课程名称','学习人数','学习时间']
        sheet_new.append(row_title)
        for i in info_dic[year]:
            row_new += 1
            sheet_new.cell(row=row_new,column=1).value = i['cre_time']
            sheet_new.cell(row=row_new,column=2).value = i['c_name']
            sheet_new.cell(row=row_new,column=3).value = i['num']
            sheet_new.cell(row=row_new,column=4).value = i['learn_time']

        wb_new.remove_sheet('Sheet')
        wb_new.save('/home/shiyanlou/Code/%s.xlsx'%year)
        wb_new.close()
#执行
if __name__ == '__main__':
    combine()
    split()

